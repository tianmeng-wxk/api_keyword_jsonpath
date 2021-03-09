import unittest,pytest,allure
from ddt import ddt, data, unpack
from common.common import excelData,Convert,depend
from key_word.http import HttpClien
import jsonpath
from log.log import Logger
from openpyxl.styles import PatternFill, Font
contents = []
if len(contents) <= 0:
    caseData = excelData().getExcel()
    '''
    {'case1.xlsx': [['http://39.98.138.157:5000/api/login', '{"username":"admin","password":"123456"}', '{"content-type": "application/json"}', 'post', 'json', 'success', '$.msg', 'loginvar'], ['http://39.98.138.157:5000/api/getproductinfo?productid=8888', None, None, 'get', 'url', 200, '$.httpstatus', 'productvar'], ['http://39.98.138.157:5000/api/getuserinfo', None, '{"token": "$loginvar.token$"}', 'get', 'url', 200, '$.httpstatus', 'uservar'], ['http://39.98.138.157:5000/api/addcart', '{"userid":$uservar.data[0].userid$,"openid":"$uservar.data[0].openid$","productid":$productvar.data[0].productid$}', '{"token": "$loginvar.token$","content-type": "application/json"}', 'post', 'json', 200, '$.httpstatus', 'cartinfo'], ['http://39.98.138.157:5000/api/createorder', '{"cartid":$cartinfo.data[0].cartid$,"openid":"$uservar.data[0].openid$","productid":$productvar.data[0].productid$,"userid":$uservar.data[0].userid$}', '{"token": "$loginvar.token$","content-type": "application/json"}', 'post', 'json', 200, '$.httpstatus', '/']],
    'case2.xlsx': [['http://39.98.138.157:5000/api/login', '{"username":"admin","password":"123456"}', '{"content-type": "application/json"}', 'post', 'json', 'success', '$.msg', 'loginvar'], ['http://39.98.138.157:5000/api/getproductinfo?productid=8888', None, None, 'get', 'url', 200, '$.httpstatus', 'productvar'], ['http://39.98.138.157:5000/api/getuserinfo', None, '{"token": "$loginvar.token$"}', 'get', 'url', 200, '$.httpstatus', 'uservar'], ['http://39.98.138.157:5000/api/addcart', '{"userid":$uservar.data[0].userid$,"openid":"$uservar.data[0].openid$","productid":$productvar.data[0].productid$}', '{"token": "$loginvar.token$","content-type": "application/json"}', 'post', 'json', 200, '$.httpstatus', 'cartinfo'], ['http://39.98.138.157:5000/api/createorder', '{"cartid":$cartinfo.data[0].cartid$,"openid":"$uservar.data[0].openid$","productid":$productvar.data[0].productid$,"userid":$uservar.data[0].userid$}', '{"token": "$loginvar.token$","content-type": "application/json"}', 'post', 'json', 200, '$.httpstatus', '/']]}
    '''
    Logger().log().info("测试数据1:{}".format(caseData))
    for name in caseData:
        listCase = caseData[name]
        Logger().log().info("测试数据2:{}".format(listCase))
        for list in listCase:
            contents.append(list)
        Logger().log().info("测试数据3:{}".format(contents))
class TestCase():
    i = 1
    @allure.title("title")
    @pytest.mark.parametrize("url, body, header, method, method_type,expect_value, jsonpath_value, dependency,status,response", contents)
    def test_collection(self, url, body, header, method, method_type,expect_value, jsonpath_value, dependency, status,response):
        Logger().log().info("开始执行第{}条用例".format(TestCase.i))
        comon=HttpClien()
        body = body.replace('\r', '').replace('\n', '').replace('\t', '') if body is not None else ""
        body = Convert().convert(body) if body.find("$") >= 0 else body


        header = Convert().convert(header) if (header is not None and header.find("$") >= 0) else header
        header = "" if header is None else header
        res=comon.request(method,url,method_type,body,header)
        #把放回数据存入字典depend相应key中
        if len(dependency)>0 and dependency.find("/")<0:
            depend[dependency] = res.content
        res = res.json()
        import openpyxl
        wb = openpyxl.load_workbook("../cases/case1.xlsx")
        ws = wb["Sheet1"]
        ws.cell(TestCase.i+1, 10).value = str(res)
        Logger().log().info("响应结果：{}".format(res))
        # 如果响应的结果不是json串，可以通过正则提取
        # import re
        # res = str(res)
        # jsonpath_value = re.findall(jsonpath_value, res)#'msg': '(.*?)'
        jsonpath_value = jsonpath.jsonpath(res,jsonpath_value)
        TestCase.i = TestCase.i + 1
        try:
        # assert expect_value == jsonpath_value[0], "断言失败"
            assert expect_value == jsonpath_value[0]
        except:
            ws.cell(TestCase.i, 9).value = "false"
            ws.cell(TestCase.i, 9).fill = PatternFill('solid', fgColor='FF0000')
            ws.cell(TestCase.i, 9).font = Font(bold=True)
            wb.save("../cases/case1.xlsx")
            assert False
        else:
            ws.cell(TestCase.i, 9).value = "pass"
            ws.cell(TestCase.i, 9).fill = PatternFill('solid', fgColor='66ff00')
            ws.cell(TestCase.i, 9).font = Font(bold=True)
            assert True
        wb.save("../cases/case1.xlsx")
        wb.close()


if __name__ == '__main__':
    pytest.main()